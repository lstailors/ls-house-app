"""Build the LS Tailors size-chart workbook (one sheet per garment type)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/ls-house-app/docs/size-charts/LS_Tailors_Size_Charts.xlsx"

FONT = "Arial"
NAVY = "1F3864"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SIZE_FILL = PatternFill("solid", fgColor="F2F2F2")
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")

thin = Side(style="thin", color="BFBFBF")
med = Side(style="medium", color=NAVY)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

FRAC = "# ?/?"   # 29.125 -> 29 1/8
CM = "0"

# ---------------------------------------------------------------- data

JACKET_COLS = ["Back Length", "Sleeve Length", "Point to Point", "Half Girth",
               "Finished Cuff", "Half Back", "Finished Bicep"]

JACKET_IN = [
    (36, 29.125, 23.625, 17.5,   18.25,  10.625, 8.375,  15.375),
    (38, 29.375, 23.75,  18.0,   19.25,  11.0,   8.625,  15.875),
    (40, 29.625, 23.875, 18.375, 20.375, 11.375, 8.875,  16.375),
    (42, 30.0,   24.125, 18.875, 21.375, 11.625, 9.125,  16.75),
    (44, 30.25,  24.375, 19.375, 22.375, 12.0,   9.375,  17.25),
    (46, 30.5,   24.625, 19.875, 23.375, 12.25,  9.625,  17.75),
    (48, 30.75,  24.75,  20.375, 24.375, 12.625, 9.875,  18.25),
    (50, 31.5,   25.0,   20.75,  25.375, 12.875, 10.125, 18.75),
]

JACKET_CM = [
    (46, 74, 60, 44, 46, 27, 21, 39),
    (48, 74, 60, 45, 49, 28, 21, 40),
    (50, 75, 60, 46, 51, 28, 22, 41),
    (52, 76, 61, 48, 54, 29, 23, 42),
    (54, 76, 61, 49, 56, 30, 23, 43),
    (56, 77, 62, 50, 59, 31, 24, 45),
    (58, 78, 62, 51, 62, 32, 25, 46),
    (60, 80, 63, 52, 64, 32, 25, 47),
]

TROUSER_COLS = ["Finished Waist", "Finished Seat", "Inseam", "Outseam",
                "1/2 Knee", "Bottom", "Finished Thigh", "Front Rise"]

TROUSER_IN = [
    (26, 26, 32.75,  31.5,   40.5,   7.25,   6.75,   22, 8.75),
    (28, 28, 34.75,  31.5,   40.5,   7.5,    6.75,   23, 8.875),
    (30, 30, 36.625, 31.5,   40.5,   7.875,  7.0,    24, 9.125),
    (32, 32, 38.5,   31.625, 40.75,  16.375, 14.375, 25, 9.375),
    (34, 34, 40.375, 31.625, 41.0,   17.125, 14.875, 26, 9.625),
    (36, 36, 42.375, 31.625, 41.25,  17.75,  15.375, 27, 9.875),
    (38, 38, 44.25,  31.625, 41.625, 18.375, 15.75,  28, 10.25),
    (40, 40, 46.125, 31.625, 41.875, 9.5,    8.0,    29, 10.5),
    (42, 42, 48.0,   31.625, 42.5,   9.875,  8.375,  30, 10.75),
    (44, 44, 50.0,   31.625, 42.5,   10.25,  8.625,  31, 11.0),
    (46, 46, 51.75,  31.75,  42.625, 10.5,   8.875,  32, 11.25),
    (48, 48, 53.625, 41.75,  43.0,   10.875, 9.125,  33, 11.625),
    (50, 50, 55.0,   31.75,  43.25,  11.25,  9.25,   34, 11.875),
]

TROUSER_CM = [
    (44,  71,  88, 80, 102, 19, 17, 58, 22),
    (46,  76,  93, 80, 102, 20, 17, 60, 23),
    (48,  81,  98, 80, 103, 21, 18, 63, 23),
    (50,  86, 102, 80, 104, 23, 19, 66, 24),
    (52,  91, 107, 80, 104, 25, 19, 68, 25),
    (54,  96, 112, 80, 105, 25, 20, 71, 26),
    (56, 101, 117, 80, 106, 25, 20, 73, 26),
    (58, 106, 121, 80, 107, 25, 21, 76, 27),
    (60, 111, 127, 80, 107, 26, 21, 78, 27),
    (62, 116, 131, 80, 108, 26, 22, 81, 28),
    (64, 121, 135, 80, 109, 27, 23, 83, 29),
    (66, 127, 139, 80, 109, 28, 23, 86, 30),
]

WAISTCOAT_COLS = ["Back Length", "Front Length", "Half Waist"]

WAISTCOAT_IN = [
    (36, 21.25,  27.25, 17.5),
    (38, 21.25,  27.75, 18.5),
    (40, 21.625, 28.25, 19.5),
    (42, 22.0,   28.75, 20.5),
    (44, 22.875, 29.25, 21.5),
    (46, 22.25,  29.75, 22.5),
    (48, 22.75,  32.25, 23.5),
    (50, 22.75,  32.75, 24.5),
]

WAISTCOAT_CM = [
    (46, 53, 69, 44),
    (48, 54, 70, 47),
    (50, 55, 71, 49),
    (52, 55, 73, 52),
    (54, 58, 74, 54),
    (56, 58, 75, 57),
    (58, 58, 81, 69),
]

# cells to highlight as "check this value" (sheet -> set of (row_size, column label))
FLAGS = {
    "Trouser": {
        ("IN", 28, "1/2 Knee"),
        ("IN", 32, "1/2 Knee"), ("IN", 34, "1/2 Knee"),
        ("IN", 36, "1/2 Knee"), ("IN", 38, "1/2 Knee"),
        ("IN", 32, "Bottom"), ("IN", 34, "Bottom"),
        ("IN", 36, "Bottom"), ("IN", 38, "Bottom"),
        ("IN", 48, "Inseam"),
    },
    "Waistcoat": {
        ("IN", 44, "Back Length"),
        ("IN", 46, "Back Length"),
        ("CM", 58, "Front Length"),
        ("CM", 58, "Half Waist"),
    },
    "Jacket-Suit-Tuxedo": set(),
}

# ---------------------------------------------------------------- helpers


def write_table(ws, start_row, title, size_header, cols, rows, numfmt, flags, unit_key):
    """Write one titled measurement table; returns the next free row."""
    ncols = len(cols) + 1
    last_col = get_column_letter(ncols)

    # title bar
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=ncols)
    c = ws.cell(row=start_row, column=1, value=title)
    c.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[start_row].height = 22

    # header row
    hr = start_row + 1
    headers = [size_header] + cols
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
        c.fill = SUB_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=med)
    ws.row_dimensions[hr].height = 30

    # data rows
    for r, row in enumerate(rows, start=hr + 1):
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = BOX
            c.alignment = Alignment(horizontal="center", vertical="center")
            if i == 1:
                c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
                c.fill = SIZE_FILL
                c.number_format = "0"
            else:
                c.font = Font(name=FONT, size=10)
                c.number_format = numfmt
                if (unit_key, row[0], cols[i - 2]) in flags:
                    c.fill = FLAG_FILL

    end = hr + len(rows)
    ws.cell(row=end + 1, column=1,
            value="Amber cells = value differs from the pattern of the column; see the "
                  "'Notes & Data Flags' tab.").font = Font(
        name=FONT, size=8, italic=True, color="808080")
    return end + 3


def style_sheet(ws, ncols):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = "B3"


# ---------------------------------------------------------------- build

wb = Workbook()

# ---- cover sheet
cover = wb.active
cover.title = "Read Me"
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 22
cover.column_dimensions["B"].width = 95

cover["A1"] = "LS Tailors — Garment Size Charts"
cover["A1"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
cover.merge_cells("A1:B1")
cover.row_dimensions[1].height = 28

lines = [
    ("Contents", "One tab per garment type. Each tab holds the Inches chart and the "
                 "Centimetres chart for that garment."),
    ("Tabs", "Jacket-Suit-Tuxedo  •  Trouser  •  Waistcoat  •  Notes & Data Flags"),
    ("Units", "Inch charts are stored as real numbers and displayed as fractions "
              "(29.125 shows as 29 1/8). Fractions display in lowest terms, so a "
              "source value of 18 2/8 shows as 18 1/4 — same measurement."),
    ("Sizing", "Inch charts use US/UK sizes (36–50). Centimetre charts use EU sizes "
               "(46–66). The two are separate published charts, not conversions of "
               "each other — small rounding differences between them are expected."),
    ("Amber cells", "A value that breaks the progression of its column. Nothing has been "
                    "changed — see the 'Notes & Data Flags' tab for what looks off and why."),
    ("Source", "Measurement values supplied by Carl (carl@lstailors.com), 3 Aug 2026."),
]
r = 3
for label, text in lines:
    c = cover.cell(row=r, column=1, value=label)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    c.alignment = Alignment(vertical="top")
    c = cover.cell(row=r, column=2, value=text)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    cover.row_dimensions[r].height = 30
    r += 2

# ---- garment sheets
specs = [
    ("Jacket-Suit-Tuxedo", "Jacket, Suit, Tuxedo", JACKET_COLS, JACKET_IN, JACKET_CM),
    ("Trouser", "Trouser", TROUSER_COLS, TROUSER_IN, TROUSER_CM),
    ("Waistcoat", "Waistcoat", WAISTCOAT_COLS, WAISTCOAT_IN, WAISTCOAT_CM),
]

for tab, label, cols, rows_in, rows_cm in specs:
    ws = wb.create_sheet(tab)
    flags = FLAGS[tab]
    nxt = write_table(ws, 1, f"{label} — Size Chart (Inches)", "Size (US/UK)",
                      cols, rows_in, FRAC, flags, "IN")
    write_table(ws, nxt, f"{label} — Size Chart (Centimetres)", "Size (EU)",
                cols, rows_cm, CM, flags, "CM")
    style_sheet(ws, len(cols) + 1)

# ---- notes sheet
nt = wb.create_sheet("Notes & Data Flags")
nt.sheet_view.showGridLines = False
for col, w in zip("ABCDE", (20, 10, 16, 16, 74)):
    nt.column_dimensions[col].width = w

nt["A1"] = "Notes & Data Flags"
nt["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
nt.merge_cells("A1:E1")

nt["A2"] = ("Every value in this workbook is exactly as supplied — nothing was corrected. "
            "These are the cells that look like transcription errors, listed so they can "
            "be checked against the master pattern block.")
nt["A2"].font = Font(name=FONT, size=10, italic=True)
nt.merge_cells("A2:E2")
nt.row_dimensions[2].height = 28
nt["A2"].alignment = Alignment(vertical="top", wrap_text=True)

note_headers = ["Chart", "Size", "Column", "Value as given", "What looks wrong"]
for i, h in enumerate(note_headers, start=1):
    c = nt.cell(row=4, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

notes = [
    ("Trouser (Inches)", 28, "1/2 Knee", "7 /12",
     "Entered as 7 1/2. The original reads '7 /12', which is not a valid fraction; 7 1/2 "
     "is the only value that fits between size 26 (7 1/4) and size 30 (7 7/8). "
     "This is the one cell where a reading had to be chosen."),
    ("Trouser (Inches)", "32–38", "1/2 Knee", "16 3/8 – 18 3/8",
     "Roughly double the neighbouring sizes (size 30 is 7 7/8, size 40 is 9 1/2). Looks "
     "like full knee width was recorded instead of half knee for these four rows."),
    ("Trouser (Inches)", "32–38", "Bottom", "14 3/8 – 15 6/8",
     "Same pattern — roughly double the neighbours (size 30 is 7, size 40 is 8). Likely "
     "full width rather than half."),
    ("Trouser (Inches)", 48, "Inseam", "41 3/4",
     "Every other size is 31 1/2 – 31 3/4. Almost certainly 31 3/4 with a mistyped first digit."),
    ("Waistcoat (Inches)", 44, "Back Length", "22 7/8",
     "Sits above size 46 (22 1/4) and size 48 (22 3/4), breaking the run. 22 1/8 would fit "
     "the progression."),
    ("Waistcoat (Inches)", 46, "Back Length", "22 1/4",
     "Flagged alongside size 44 — one of the two is out of sequence."),
    ("Waistcoat (Cm)", 58, "Front Length", "81",
     "Jumps 6 cm from size 56 (75) where every other step is 1–2 cm."),
    ("Waistcoat (Cm)", 58, "Half Waist", "69",
     "Jumps 12 cm from size 56 (57) where every other step is 2–3 cm. 59 or 60 would fit."),
    ("Waistcoat (Cm)", "60", "(whole row)", "missing",
     "The cm chart stops at size 58 while the inch chart runs to size 50 (8 rows vs 7). "
     "A size 60 row may be missing."),
]

for r, note in enumerate(notes, start=5):
    for i, val in enumerate(note, start=1):
        c = nt.cell(row=r, column=i, value=val)
        c.font = Font(name=FONT, size=10)
        c.border = BOX
        c.alignment = Alignment(horizontal="center" if i in (2, 3, 4) else "left",
                                vertical="top", wrap_text=True)
        if i == 4:
            c.fill = FLAG_FILL
    nt.row_dimensions[r].height = 42

wb.save(OUT)
print("wrote", OUT)
